from pathlib import Path
from typing import Optional

try:
    import pypdf  # type: ignore
except ModuleNotFoundError:  # pragma: no cover
    try:
        import PyPDF2 as pypdf  # type: ignore
    except ModuleNotFoundError:  # pragma: no cover
        pypdf = None  # type: ignore[assignment]

try:
    from docx import Document as DocxDocument  # type: ignore
except ModuleNotFoundError:  # pragma: no cover
    DocxDocument = None  # type: ignore[assignment]

try:
    from openpyxl import load_workbook  # type: ignore
except ModuleNotFoundError:  # pragma: no cover
    load_workbook = None  # type: ignore[assignment]

SUPPORTED_TYPES = {"pdf", "doc", "docx", "md", "xls", "xlsx"}
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB


class DocumentParser:
    @staticmethod
    def parse(file_path: Path, file_type: str) -> tuple[str, bool]:
        """Parse document and extract text. Returns (text, is_encrypted)."""
        try:
            file_type = file_type.lower()

            if file_type == "pdf":
                return DocumentParser._parse_pdf(file_path)
            elif file_type in ("doc", "docx"):
                return DocumentParser._parse_docx(file_path)
            elif file_type in ("xls", "xlsx"):
                return DocumentParser._parse_excel(file_path)
            elif file_type == "md":
                return DocumentParser._parse_markdown(file_path)
            return "", False
        except Exception:
            return "", False

    @staticmethod
    def _parse_pdf(file_path: Path) -> tuple[str, bool]:
        if pypdf is None:
            return "", False
        reader = pypdf.PdfReader(str(file_path))
        if reader.is_encrypted:
            return "", True
        text = ""
        for page in reader.pages:
            text += page.extract_text() or ""
        return text, False

    @staticmethod
    def _parse_docx(file_path: Path) -> tuple[str, bool]:
        if DocxDocument is None:
            return "", False
        doc = DocxDocument(str(file_path))
        text = "\n".join([p.text for p in doc.paragraphs])
        return text, False

    @staticmethod
    def _parse_excel(file_path: Path) -> tuple[str, bool]:
        if load_workbook is None:
            return "", False
        wb = load_workbook(str(file_path), read_only=True, data_only=True)
        text_parts = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                row_text = " ".join(str(cell) for cell in row if cell is not None)
                if row_text.strip():
                    text_parts.append(row_text)
        return "\n".join(text_parts), False

    @staticmethod
    def _parse_markdown(file_path: Path) -> tuple[str, bool]:
        with open(file_path, "r", encoding="utf-8") as f:
            return f.read(), False

    @staticmethod
    def is_supported(file_type: str) -> bool:
        return file_type.lower() in SUPPORTED_TYPES

    @staticmethod
    def get_file_type(filename: str) -> str:
        return Path(filename).suffix.lower().lstrip(".")
